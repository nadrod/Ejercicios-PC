# GALVAN GARCIA GILBERTO EDUARDO, GAYTAN MONTELONGO JESUS GERARDO
# DIAZ PINEDA LUIS ROBERTO, PEREZ HERNANDEZ JOSE PABLO

from pyhunter import PyHunter
from openpyxl import Workbook
from openpyxl import load_workbook
import getpass
import re

correo = re.compile(r'([a-zA-z]+).([a-zA-z]+)@(\D+)(.com|.mx)')
redes = re.compile(
                   r'(https://www.|https://)/'
                   r'(facebook|instagram|youtube|twitter)/'
                   r'(.com/user/|.com/)?([a-zA-z]+)')
dominio = re.compile(r'([a-zA-z]+)(.com|.mx)')


def Busqueda(organizacion):
    # Cantidad de resultados esperados de la búsqueda
    # El límite MENSUAL de Hunter es 50, cuidado!
    resultado = hunter.domain_search(company=orga,
                                     limit=1, emails_type='personal')
    return resultado


def GuardarInformacion(datosEncontrados, orga):
    try:
        libro = load_workbook("Hunter" + orga + ".xlsx")
    except FileNotFoundError:
        libro = Workbook()

    # seleccionar hoja
    if len(libro.sheetnames) > 1:
        libro.active = 1
        hoja = libro.active
    else:
        hoja = libro.active
    hoja["A1"] = "Organizacion"
    hoja["B1"] = "Redes"
    hoja["C1"] = "Correos"
    hoja["D1"] = "Dominios"
    hoja["A2"] = orga
    libro.save("Hunter" + orga + ".xlsx")

    linksfa = []
    linksfauno = []
    with open("archivo.txt", "r", encoding='UTF-8') as file2:
        for linea2 in file2:
            moredesfa = redes.search(str(linea2))
            if (moredesfa is not None):
                # print(moredesfa.group())
                linksfa.append(moredesfa.group())
                linksfauno = list(dict.fromkeys(linksfa))

    for nfa in range(len(linksfauno)):
        hoja["B"+str(nfa + 2)] = str(linksfauno[nfa])

    libro.save("Hunter" + orga + ".xlsx")

    linksco = []
    linkscouno = []
    with open("archivo.txt", "r", encoding='UTF-8') as file2:
        for linea2 in file2:
            moredesco = correo.search(str(linea2))
            if (moredesco is not None):
                # print(moredesfa.group())
                linksco.append(moredesco.group())
                linkscouno = list(dict.fromkeys(linksco))

    for nfa in range(len(linkscouno)):
        hoja["C"+str(nfa + 2)] = str(linkscouno[nfa])

    libro.save("Hunter" + orga + ".xlsx")

    linksdo = []
    linksdouno = []
    with open("archivo.txt", "r", encoding='UTF-8') as file2:
        for linea2 in file2:
            moredesdo = dominio.search(str(linea2))
            if (moredesdo is not None):
                # print(moredesfa.group())
                linksdo.append(moredesdo.group())
                linksdouno = list(dict.fromkeys(linksdo))

    for nfa in range(len(linksdouno)):
        hoja["D"+str(nfa + 2)] = str(linksdouno[nfa])

    libro.save("Hunter" + orga + ".xlsx")
    print("Informacion guardada")

print("Script para buscar información")
apikey = getpass.getpass("Ingresa tu API key: ")
hunter = PyHunter(apikey)
orga = input("Dominio a investigar: ")
datosEncontrados = Busqueda(orga)
archivo = open("archivo.txt", "w")
a = datosEncontrados
archivo.write(str((a)))
archivo.close()
if datosEncontrados is None:
    exit()
else:
    GuardarInformacion(datosEncontrados, orga)
